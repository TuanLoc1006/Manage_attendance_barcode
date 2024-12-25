import sys
from PyQt5.QtWidgets import QApplication,QMainWindow, QFileDialog,QMessageBox 
from quanLyDiemDanh import Ui_MainWindow
from PyQt5.QtGui import QPixmap
from PyQt5 import QtGui,QtWidgets
from barcode import Code128
from io import BytesIO
from barcode.writer import ImageWriter
rv = BytesIO()

import xuLyDangKy
# mo webcam
import cv2
from pyzbar.pyzbar import decode
import numpy as np 

# xuat du lieu ra file
import openpyxl
# tao am thanh
import winsound
frequency = 2000  # Set Frequency To 2500 Hertz
duration = 500  # Set Duration To 1000 ms == 1 second
# import ham time
import time
#ket noi voi xampp
import mysql.connector
# import ngay gio
import datetime

mydb = mysql.connector.connect(
    host='localhost',
    username='root',
    password='',
    port='3306',
    database='project_barcode'
)

class MainWindow:
    def __init__(self):
        self.main_win = QMainWindow()
        self.uic = Ui_MainWindow()
        self.uic.setupUi(self.main_win)

        self.uic.btn_openCam.clicked.connect(self.openCameraBarCode)
        self.uic.btn_export_file.clicked.connect(self.export_to_excel)
        self.uic.btn_tim_danhSach.clicked.connect(self.layDanhSachLop)
        self.uic.btn_luu_danh_sach.clicked.connect(self.luu_danh_sach)

        self.uic.btn_page_dangky.clicked.connect(self.pageDangky)
        self.uic.btn_closeCam.clicked.connect(self.closeCam)
        self.uic.btn_clear_input.clicked.connect(self.clear_input)

        # Lấy ngày hiện tại
        # today = datetime.date.today()
        # Chuyển đổi sang ngày tháng năm
        # day = today.strftime("%d/%m/%Y")
        # self.uic.input_ngay.setText(day)

    def pageDangky(self):
        self.main_win = xuLyDangKy.MainWindow()
        self.main_win.show()

    def closeCam(self):
        self.cam_0 = cv2.VideoCapture(0) 
        self.cam_0.release()
        self.clearData()

    def clear_input(self):
        self.uic.input_ma_mon.setText("")
        self.uic.input_nhom.setText("")
        self.uic.input_tuan_hoc.setText("")
        self.clearData()
       
    def openCameraBarCode(self):
        self.cam_0 = cv2.VideoCapture(0) 
        while True:
            try:
                _, self.frame_so_0 = self.cam_0.read()

                # xu ly nhan dien ma vach
                for code in decode(self.frame_so_0):
                    pts = np.array([code.polygon], np.int32)
                    pts = pts.reshape(((-1,1,2)))
                    cv2.polylines(self.frame_so_0, [pts], True, (0,0,255), 3)
                    data = code.data.decode('utf-8')
                    winsound.Beep(frequency, duration)
                    time.sleep(1)   
                    self.handle_barcode_scanned(data)
                    self.getInformation(data)
                # hien len man hinh fram
                self.frame_so_0 = cv2.cvtColor(self.frame_so_0, cv2.COLOR_BGR2RGB)
                self.update()
                
                if cv2.waitKey(1) & 0xFF == ord('q'):
                    self.cam_0.release()
                    cv2.destroyAllWindows()
                    break
            except IndexError:
                self.clearData()
                msg = QMessageBox()
                msg.setText('Ma vach khong dung, xin vui long quet lai!')
                msg.exec_()
                
            except TypeError:
                break



    def update(self):
        self.setPhoto(self.frame_so_0)
        

    def setPhoto(self,image):
        image = cv2.resize(image,(331,251))
        img = QtGui.QImage(image,image.shape[1],image.shape[0], image.strides[0],QtGui.QImage.Format_RGB888)
        self.uic.label_frameBarCode.setPixmap(QtGui.QPixmap.fromImage(img))

    def clearData(self):
        self.uic.input_hoten.setText("")
        self.uic.input_email.setText("")
        self.uic.input_mssv.setText("")
        self.uic.input_sdt.setText("")
        
        self.uic.label_sinh_vien.setPixmap(QPixmap("E:\\CTU\\HK2_nam_4_2023-2024\\NLMMT_full\\img\\pic_user.png"))
    def getInformation(self, mssv):
        mycursor = mydb.cursor()
        sql = " SELECT * FROM `sinh_vien` WHERE ma_so_sinh_vien = %s"
        adr = (mssv,)
        mycursor.execute(sql, adr)
        myresult = mycursor.fetchall()
        print(myresult)
        self.uic.input_mssv.setText(myresult[0][0])     
        self.uic.input_hoten.setText(myresult[0][1])
        self.uic.input_email.setText(myresult[0][2]) 
        self.uic.input_sdt.setText(myresult[0][4]) 
        self.uic.label_sinh_vien.setPixmap(QPixmap(myresult[0][5]))
        
    def handle_barcode_scanned(self, barcode_data):
        try:
            tuan_diem_danh = self.uic.input_tuan_hoc.text()
            # print('tuan dang diem danh: ' + tuan_diem_danh)
            found = False
            for row in range(self.uic.bang_danh_sach.rowCount()):
                item = self.uic.bang_danh_sach.item(row, 0)  # Giả sử mã vạch ở cột đầu tiên
                if item is not None and item.text() == barcode_data:
                    # Xác định cột tương ứng với tuan_diem_danh
                    current_week_column = int(tuan_diem_danh)  
                    # Cập nhật trạng thái điểm danh (cột tương ứng với tuan_diem_danh)
                    status_item = QtWidgets.QTableWidgetItem("Có mặt")
                    self.uic.bang_danh_sach.setItem(row, current_week_column, status_item)
                    found = True
                    break

            if not found:
                # Nếu mã vạch không khớp với dữ liệu nào, thêm hàng mới vào bảng
                # row_count = self.uic.bang_danh_sach.rowCount()
                # self.uic.bang_danh_sach.insertRow(row_count)
                # # Đặt mã vạch vào cột đầu tiên
                # barcode_item = QtWidgets.QTableWidgetItem(barcode_data)
                # self.uic.bang_danh_sach.setItem(row_count, 0, barcode_item)
                # # Đặt trạng thái mặc định vào cột tương ứng với tuan_diem_danh
                # current_week_column = int(tuan_diem_danh)   
                # status_item = QtWidgets.QTableWidgetItem("No details")
                # self.uic.bang_danh_sach.setItem(row_count, current_week_column, status_item)
                return

        except Exception as e:
            message = QMessageBox()
            message.setText(f"Vui lòng chọn mã học phần và nhóm")
            message.exec_()


    def layDanhSachLop(self):
        try:
            ma_hoc_phan = self.uic.input_ma_mon.text()
            nhom = self.uic.input_nhom.text()
            mycursor = mydb.cursor()
            today = datetime.date.today()
            ngay = today.strftime("%d/%m/%Y")
            tuan_diem_danh = self.uic.input_tuan_hoc.text()
            if tuan_diem_danh == "":
                message = QtWidgets.QMessageBox()
                message.setText("vui lòng thêm thông tin tuần học")
                message.exec_()
                return
            # Thực hiện truy vấn SQL để lấy danh sách mssv đăng ký môn học theo học phần và nhóm
            mycursor.execute("""
                SELECT dk.mssv
                FROM `dang_ky_mon_hoc` dk 
                WHERE dk.ma_mon = %s and dk.nhom= %s
            """, (ma_hoc_phan, nhom))
            myresult = mycursor.fetchall()
            print(myresult)
            # Xóa dữ liệu cũ trong bảng
            self.uic.bang_danh_sach.setRowCount(0)

            # Thêm dữ liệu mới vào bảng
            if myresult:
                # Tạo một danh sách các mssv từ kết quả truy vấn
                mssv_list = [row[0] for row in myresult]

                num_rows = len(mssv_list)
                num_cols = 21  # Số cột mới từ tuần 1 đến tuần 15

                # Thiết lập số hàng và số cột trong bảng
                self.uic.bang_danh_sach.setRowCount(num_rows)
                self.uic.bang_danh_sach.setColumnCount(num_cols)

                # Đặt tiêu đề cho cột mssv
                self.uic.bang_danh_sach.setHorizontalHeaderItem(0, QtWidgets.QTableWidgetItem('mssv'))

                # Đặt tiêu đề cho các cột từ tuần 1 đến tuần 21
                for col_number in range(1, num_cols):
                    tuan_title = f"tuần {col_number}"
               
                    self.uic.bang_danh_sach.setHorizontalHeaderItem(col_number, QtWidgets.QTableWidgetItem(tuan_title))
                
                
                # Thêm dữ liệu vào bảng             
                for row_number, mssv in enumerate(mssv_list):
                    # Đặt mssv trong cột đầu tiên
                    self.uic.bang_danh_sach.setItem(row_number, 0, QtWidgets.QTableWidgetItem(str(mssv)))

                    # Lặp qua từ tuần 1 đến tuần 20 để lấy thông tin trạng thái điểm danh
                    for col_number in range(1, num_cols):
                        # Thực hiện truy vấn SQL để lấy thông tin trạng thái
                        mycursor.execute("""
                            SELECT trang_thai
                            FROM `diem_danh`
                            WHERE mssv = %s AND nhom= %s AND ma_mon = %s AND tuan_hoc = %s
                        """, (mssv,nhom, ma_hoc_phan, col_number))

                        # Lấy kết quả truy vấn
                        trang_thai_result = mycursor.fetchone()

                        # Kiểm tra xem có kết quả không
                        if trang_thai_result:
                            trang_thai = trang_thai_result[0]
                            self.uic.bang_danh_sach.setItem(row_number, col_number, QtWidgets.QTableWidgetItem(trang_thai))
                        else:
                            # Nếu không có kết quả, đặt giá trị rỗng
                            self.uic.bang_danh_sach.setItem(row_number, col_number, QtWidgets.QTableWidgetItem(''))
            else:
                message = QtWidgets.QMessageBox()
                message.setText("không tìm thấy học phần, vui lòng kiểm tra lại mã học phần và nhóm!")
                message.exec_()

        except Exception as e:
            # Xử lý ngoại lệ nếu cần thiết
            message = QtWidgets.QMessageBox()
            message.setText(f"Lỗi: {str(e)}")
            message.exec_()

    def set_column_background_color(self, column, color):
        # Đặt màu nền cho nguyên cột
        for row in range(self.uic.bang_danh_sach.rowCount()):
            item = self.uic.bang_danh_sach.item(row, column)
            if item:
                item.setBackground(QtGui.QBrush(color))
   


    def luu_danh_sach(self):
        try:
            ma_hoc_phan = self.uic.input_ma_mon.text()
            nhom = self.uic.input_nhom.text()
            tuan = self.uic.input_tuan_hoc.text()
            today = datetime.date.today()
            ngay = today.strftime("%d/%m/%Y")
          
            mycursor = mydb.cursor()

            for row in range(self.uic.bang_danh_sach.rowCount()):
                mssv_item = self.uic.bang_danh_sach.item(row, 0)  # Cột mã số sinh viên

                # Lặp qua từng cột từ tuần 1 đến tuần cuối cùng
                for col_number in range(int(tuan), self.uic.bang_danh_sach.columnCount()):
                    trang_thai_item = self.uic.bang_danh_sach.item(row, col_number)  # Cột trạng thái điểm danh
                    # print(type(col_number), type(tuan))
                    if mssv_item is not None and trang_thai_item is not None :
                        mssv = mssv_item.text()
                        trang_thai = trang_thai_item.text()

                        # Kiểm tra xem trạng thái đã được điểm danh hay chưa
                        if trang_thai == "Có mặt" or trang_thai == "No details":

                            # sql_tkb = "INSERT INTO `thoi_khoa_bieu`(`tuan_hoc`, `ngay`) VALUES (%s,%s)"
                            # val_tkb = (tuan, str(ngay))
                            # mycursor.execute(sql_tkb, val_tkb)
                            # mydb.commit()
                            # Lưu vào CSDL
                            sql_diem_danh = " INSERT INTO `diem_danh`(`id`, `mssv`, `ma_mon`, `nhom`, `tuan_hoc`, `trang_thai`) VALUES ('', %s, %s, %s, %s, %s)"
                            val_diem_danh = (mssv, ma_hoc_phan, nhom, tuan, trang_thai)
                            print(mssv, ma_hoc_phan,'nhom:'+ nhom, 'tuan:'+tuan, trang_thai, ngay)
                            mycursor.execute(sql_diem_danh, val_diem_danh)

                        
            mydb.commit()
            success_message = QMessageBox()
            success_message.setText("Lưu điểm danh thành công")
            success_message.exec_()

        except Exception as e:
            error_message = QMessageBox()
            error_message.setText(f"Lỗi: {str(e)}")
            print(e)
            error_message.exec_()


    def export_to_excel(self):
       
        ma_mon = self.uic.input_ma_mon.text()
        nhom = self.uic.input_nhom.text()
        # ngay = self.uic.input_ngay.text()
        print(ma_mon,nhom)
        # Workbook and active sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        sheet['A1'] = 'Ma mon'
        sheet['B1'] = ma_mon
        
        sheet['C1'] = 'nhom'
        sheet['D1'] = nhom
        # Set column headers from the table
        col_count = self.uic.bang_danh_sach.columnCount()
        for col in range(col_count):
            header_item = self.uic.bang_danh_sach.horizontalHeaderItem(col)
            if header_item is not None:
                sheet.cell(row=2, column=col + 1).value = header_item.text()

        # Set data from the table
        row_count = self.uic.bang_danh_sach.rowCount()
        for row in range(row_count):
            for col in range(col_count):
                item = self.uic.bang_danh_sach.item(row, col)
                if item is not None:
                    sheet.cell(row=row + 3, column=col + 1).value = item.text()

        # Mở hộp thoại lưu file
        file_dialog = QFileDialog()
        file_name, _ = file_dialog.getSaveFileName(self.main_win, 'Luu file Excel', 'danh_sach_diem_danh.xlsx', 'Excel Files (*.xlsx)')

        if not file_name:
            # Người dùng nhấn Cancel hoặc không nhập tên file
            return

        try:
            workbook.save(file_name)
            message = QMessageBox()
            message.setText("Xuất file Excel thành công!")
            message.exec_()
        except Exception as e:
            message = QMessageBox()
            message.setText(f"Lỗi khi xuất file Excel: {e}")
            message.exec_()

            
    def show(self):
        self.main_win.show()

# if __name__ == "__main__":
#     app = QApplication(sys.argv)
#     main_win = MainWindow()
#     main_win.show()
#     sys.exit(app.exec())